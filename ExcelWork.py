from openpyxl import load_workbook
from datetime import datetime


# Load the workbook
filename = "data.xlsx"
workbook = load_workbook(filename)

# Select the active sheet
sheet = workbook.active

# Determine the next available row
next_row = sheet.max_row + 1

# Write data to the sheet
def saveData(user):
    name = user
    email = "null"
    if name == "Pasindu":
        email = "kvpasindumalinda@gmail.com"
    elif name == "Lakindu":
        email = "lakindubannaheka@gmail.com"

    status = "Door Unlocked Successfully"
    now = datetime.now()
    current_time = now.strftime("%H-%M-%S")
    current_date = now.strftime("%Y-%m-%d")
    sheet.cell(row=next_row, column=1).value = name
    sheet.cell(row=next_row, column=2).value = email
    sheet.cell(row=next_row, column=3).value = status
    sheet.cell(row=next_row, column=4).value = current_time
    sheet.cell(row=next_row, column=5).value = current_date

    # Save the workbook
    workbook.save(filename)
    print("Data appended successfully.")